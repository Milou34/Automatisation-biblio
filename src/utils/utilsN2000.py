from openpyxl.styles import Font, Alignment

def create_table_hab_n2000(ws, current_row):
    """
    Crée un en-tête de tableau avec un titre en gras et centré avant, et des colonnes fusionnées.
    La fusion commence à partir de la colonne A.
    
    :param ws: La feuille de calcul (Worksheet)
    :param current_row: La ligne actuelle où l'en-tête doit être inséré
    """
    
    # Ajouter le titre principal et le fusionner de A à K (colonnes 1 à 11)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=11)
    title_cell = ws.cell(row=current_row, column=1)
    title_cell.value = "Types d’habitats présents sur le site et évaluations"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Passer à la ligne suivante pour l'en-tête des colonnes
    current_row += 1
    
    # Fusionner les cellules de A à G sur la première ligne (fusion horizontale)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
    title_cell = ws.cell(row=current_row, column=1)
    title_cell.value = "Types d’habitats inscrits à l’annexe I"
    title_cell.font = Font(bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Définir les titres pour chaque colonne de A à G
    titles = ["Code", "Habitat", "PF", "Superficie (Ha)", "Superficie (% de couverture)", "Grottes (nombre)", "Qualité des données"]

    # Fusionner les cellules de A à G sur les deux lignes suivantes (fusion verticale) et utiliser les noms fournis
    for col in range(1, 8):  # Colonnes de A (1) à G (7)
        ws.merge_cells(start_row=current_row + 1, start_column=col, end_row=current_row + 2, end_column=col)
        cell = ws.cell(row=current_row + 1, column=col)
        cell.value = titles[col - 1]  # Utiliser les noms de la liste
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Fusionner les cellules de H à K sur la première ligne (fusion horizontale)
    ws.merge_cells(start_row=current_row, start_column=8, end_row=current_row, end_column=11)
    title_cell_G_J = ws.cell(row=current_row, column=8)
    title_cell_G_J.value = "Évaluation du site"
    title_cell_G_J.font = Font(bold=True)
    title_cell_G_J.alignment = Alignment(horizontal="center", vertical="center")

    # Fusionner les cellules de la colonne H sur les deux lignes suivantes (fusion verticale)
    ws.merge_cells(start_row=current_row + 1, start_column=8, end_row=current_row + 2, end_column=8)
    cell_G = ws.cell(row=current_row + 1, column=8)
    cell_G.value = "Représentativité (A|B|C|D)"
    cell_G.font = Font(bold=True)
    cell_G.alignment = Alignment(horizontal="center", vertical="center")

    # Fusionner les cellules des colonnes I à K sur la deuxième ligne (fusion horizontale)
    ws.merge_cells(start_row=current_row + 1, start_column=9, end_row=current_row + 1, end_column=11)
    title_H_J = ws.cell(row=current_row + 1, column=9)
    title_H_J.value = "A|B|C"
    title_H_J.font = Font(bold=True)
    title_H_J.alignment = Alignment(horizontal="center", vertical="center")

    titles = ["Superficie relative", "Conservation", "Évaluation globale"]
    # Ajouter des titres dans les cellules I2 à K2 (troisième ligne)
    for col in range(9, 12):  # Colonnes I (9) à K (11)
        ws.cell(row=current_row + 2, column=col, value=titles[col - 9]).font = Font(bold=True)
        ws.cell(row=current_row + 2, column=col).alignment = Alignment(horizontal="center", vertical="center")

    # Retourner la nouvelle ligne pour la suite
    return current_row + 3



def create_table_especes_inscrites(ws, current_row):
    """Crée un en-tête de tableau avec un titre en gras et centré, et des colonnes fusionnées.
    
    :param ws: La feuille de calcul (Worksheet)
    :param current_row: La ligne actuelle où l'en-tête doit être inséré
    """
    
    # Ajouter le titre principal et le fusionner de A à J (colonnes 1 à 10)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
    title_cell = ws.cell(row=current_row, column=1)
    title_cell.value = "Espèces inscrites à l’annexe II de la directive 92/43/CEE et évaluation"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Passer à la ligne suivante pour l'en-tête des colonnes
    current_row += 1
    
    # Fusionner les cellules de A à C sur la première ligne (fusion horizontale)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    title_cell = ws.cell(row=current_row, column=1)
    title_cell.value = "Espèce"
    title_cell.font = Font(bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    titles = ["Groupe", "Code", "Nom scientifique"]
    # Fusionner les cellules de A à C sur les deux lignes suivantes (fusion verticale) et utiliser les noms fournis
    for col in range(1, 4):  # Colonnes A (1) à C (3)
        ws.merge_cells(start_row=current_row + 1, start_column=col, end_row=current_row + 2, end_column=col)
        cell = ws.cell(row=current_row + 1, column=col)
        cell.value = titles[col - 1]  # Utiliser les noms de la liste
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Fusionner les cellules de D à I sur la première ligne (fusion horizontale)
    ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=9)
    title_cell_D_I = ws.cell(row=current_row, column=4)
    title_cell_D_I.value = "Population présente sur le site"
    title_cell_D_I.font = Font(bold=True)
    title_cell_D_I.alignment = Alignment(horizontal="center", vertical="center")

    # Fusionner la colonne D verticalement
    ws.merge_cells(start_row=current_row + 1, start_column=4, end_row=current_row + 2, end_column=4)
    cell_D = ws.cell(row=current_row + 1, column=4)
    cell_D.value = "Type"
    cell_D.font = Font(bold=True)
    cell_D.alignment = Alignment(horizontal="center", vertical="center")

    # Fusionner les colonnes E à F sur la deuxième ligne (fusion horizontale)
    ws.merge_cells(start_row=current_row + 1, start_column=5, end_row=current_row + 1, end_column=6)
    title_E_F = ws.cell(row=current_row + 1, column=5)
    title_E_F.value = "Taille"
    title_E_F.font = Font(bold=True)
    title_E_F.alignment = Alignment(horizontal="center", vertical="center")

    # Titres dans la troisième ligne des colonnes E et F
    sub_titles_E_F = ["Min", "Max"]
    for col in range(5, 7):  # Colonnes E (5) et F (6)
        ws.cell(row=current_row + 2, column=col, value=sub_titles_E_F[col - 5]).font = Font(bold=True)
        ws.cell(row=current_row + 2, column=col).alignment = Alignment(horizontal="center")

    # Fusion verticale dans les colonnes G à J
    sub_titles_G_J = ["Unité", "Catégorie (C|R|V|P)", "Qualité des données", "Population (A|B|C|D)"]
    for col in range(7, 11):  # Colonnes G (7) à J (10)
        ws.merge_cells(start_row=current_row + 1, start_column=col, end_row=current_row + 2, end_column=col)
        cell_G = ws.cell(row=current_row + 1, column=col)
        cell_G.value = sub_titles_G_J[col - 7]  # Utiliser les noms de la liste
        cell_G.font = Font(bold=True)
        cell_G.alignment = Alignment(horizontal="center", vertical="center")

    # Fusion horizontale des colonnes J à M
    ws.merge_cells(start_row=current_row, start_column=10, end_row=current_row, end_column=13)
    title_K_M = ws.cell(row=current_row, column=10)
    title_K_M.value = "Évaluation du site"
    title_K_M.font = Font(bold=True)
    title_K_M.alignment = Alignment(horizontal="center", vertical="center")

    # Fusion horizontale sur la deuxième ligne des colonnes K à M
    ws.merge_cells(start_row=current_row + 1, start_column=11, end_row=current_row + 1, end_column=13)
    title_K_M_2 = ws.cell(row=current_row + 1, column=11)
    title_K_M_2.value = "A|B|C"
    title_K_M_2.font = Font(bold=True)
    title_K_M_2.alignment = Alignment(horizontal="center", vertical="center")

    # Titres dans la troisième ligne des colonnes K à M
    sub_titles_K_M = ["Conservation", "Isolement", "Évaluation globale"]
    for col in range(11, 14):  # Colonnes K (11) à M (13)
        ws.cell(row=current_row + 2, column=col, value=sub_titles_K_M[col - 11]).font = Font(bold=True)
        ws.cell(row=current_row + 2, column=col).alignment = Alignment(horizontal="center")

    # Retourner la nouvelle ligne pour la suite
    return current_row + 3


def create_table_especes_autres(ws, current_row):
    """
    Crée un en-tête de tableau avec des colonnes fusionnées et des titres en gras et centrés.
    
    :param ws: La feuille de calcul (Worksheet)
    :param current_row: La ligne actuelle où l'en-tête doit être inséré
    """
    
    # Ajouter le titre principal du tableau et le fusionner de A à L (colonnes 1 à 12)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=12)
    title_cell = ws.cell(row=current_row, column=1)
    title_cell.value = "Autres espèces importantes de faune et de flore"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Passer à la ligne suivante pour l'en-tête des colonnes
    current_row += 1

    # Fusionner les cellules de A à B sur la première ligne
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    title_cell_A_B = ws.cell(row=current_row, column=1)
    title_cell_A_B.value = "Espèce"
    title_cell_A_B.font = Font(bold=True)
    title_cell_A_B.alignment = Alignment(horizontal="center", vertical="center")

    # Fusionner les colonnes A et B verticalement sur les deux lignes suivantes (ligne 2 et 3)
    titles_A_B = ["Groupe", "Nom scientifique"]
    for col in range(1, 3):  # Colonnes A (1) et B (2)
        ws.merge_cells(start_row=current_row + 1, start_column=col, end_row=current_row + 2, end_column=col)
        cell = ws.cell(row=current_row + 1, column=col)
        cell.value = titles_A_B[col - 1]
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Fusionner les cellules de C à F sur la première ligne
    ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=6)
    title_cell_C_F = ws.cell(row=current_row, column=3)
    title_cell_C_F.value = "Population présente sur le site"
    title_cell_C_F.font = Font(bold=True)
    title_cell_C_F.alignment = Alignment(horizontal="center", vertical="center")

    # Fusionner les colonnes C et D horizontalement sur la deuxième ligne
    ws.merge_cells(start_row=current_row + 1, start_column=3, end_row=current_row + 1, end_column=4)
    title_cell_C_D = ws.cell(row=current_row + 1, column=3)
    title_cell_C_D.value = "Taille"
    title_cell_C_D.font = Font(bold=True)
    title_cell_C_D.alignment = Alignment(horizontal="center", vertical="center")

    # Titres dans la troisième ligne des colonnes C et D
    sub_titles_C_D = ["Min", "Max"]
    for col in range(3, 5):  # Colonnes C (3) et D (4)
        ws.cell(row=current_row + 2, column=col, value=sub_titles_C_D[col - 3]).font = Font(bold=True)
        ws.cell(row=current_row + 2, column=col).alignment = Alignment(horizontal="center", vertical="center")

    # Fusion verticale des colonnes E et F sur les deux lignes suivantes
    sub_titles_E_F = ["Unité", "Catégorie (C|R|V|P)"]
    for col in range(5, 7):  # Colonnes E (5) et F (6)
        ws.merge_cells(start_row=current_row + 1, start_column=col, end_row=current_row + 2, end_column=col)
        cell = ws.cell(row=current_row + 1, column=col)
        cell.value = sub_titles_E_F[col - 5]
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Fusionner les colonnes G à L sur la première ligne
    ws.merge_cells(start_row=current_row, start_column=7, end_row=current_row, end_column=12)
    title_cell_G_L = ws.cell(row=current_row, column=7)
    title_cell_G_L.value = "Motivation"
    title_cell_G_L.font = Font(bold=True)
    title_cell_G_L.alignment = Alignment(horizontal="center", vertical="center")

    # Fusionner les colonnes G et H horizontalement sur la deuxième ligne
    ws.merge_cells(start_row=current_row + 1, start_column=7, end_row=current_row + 1, end_column=8)
    title_cell_G_H = ws.cell(row=current_row + 1, column=7)
    title_cell_G_H.value = "Annexe Directive Habitat"
    title_cell_G_H.font = Font(bold=True)
    title_cell_G_H.alignment = Alignment(horizontal="center", vertical="center")

    # Titres dans la troisième ligne des colonnes G et H
    sub_titles_G_H = ["IV", "V"]
    for col in range(7, 9):  # Colonnes G (7) et H (8)
        ws.cell(row=current_row + 2, column=col, value=sub_titles_G_H[col - 7]).font = Font(bold=True)
        ws.cell(row=current_row + 2, column=col).alignment = Alignment(horizontal="center", vertical="center")

    # Fusionner les colonnes I à L horizontalement sur la deuxième ligne
    ws.merge_cells(start_row=current_row + 1, start_column=9, end_row=current_row + 1, end_column=12)
    title_cell_I_L = ws.cell(row=current_row + 1, column=9)
    title_cell_I_L.value = "Autres catégories"
    title_cell_I_L.font = Font(bold=True)
    title_cell_I_L.alignment = Alignment(horizontal="center", vertical="center")

    # Titres dans la troisième ligne des colonnes I à L
    sub_titles_I_L = ["A", "B", "C", "D"]
    for col in range(9, 13):  # Colonnes I (9) à L (12)
        ws.cell(row=current_row + 2, column=col, value=sub_titles_I_L[col - 9]).font = Font(bold=True)
        ws.cell(row=current_row + 2, column=col).alignment = Alignment(horizontal="center", vertical="center")

    # Retourner la nouvelle ligne pour la suite
    return current_row + 3