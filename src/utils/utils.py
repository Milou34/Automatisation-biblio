from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import psutil
import os
import re

INT_PATTERN = r"^-?\d+$"
FLOAT_PATTERN = r"^-?\d+\,\d+$"
FLOAT_PATTERN2 = r"^-?\,\d+$"

def extract_info(root, tag_paths):
    """Extrait les valeurs des balises spécifiées depuis le XML et retourne une liste des valeurs en string.

    Args:
        root (ET.Element): La racine du document XML.
        tag_paths (list of str): Liste des chemins de balises à extraire.

    Returns:
        list of str: Liste des valeurs extraites sous forme de chaînes de caractères.
    """
    values = []
    
    for path in tag_paths:
        element = root.find(path)
        element = element.text if element is not None and element.text is not None else "-"
        if re.match(INT_PATTERN, element): 
            # Ajoute la valeur de l'élément ou une chaîne vide si l'élément est None
            values.append(int(element))
        elif re.match(FLOAT_PATTERN, element) or re.match(FLOAT_PATTERN2, element):
            element = float((element.replace(",", ".")))
            values.append(element)
        else:
            values.append(element)
            
    return values


def create_table(ws, title, headers, start_row):
    """Crée un tableau avec un titre et des en-têtes stylisés dans la feuille de calcul."""
    # Ajouter le titre du tableau et le mettre en gras
    ws.append([title])
    title_cell = ws.cell(row=start_row, column=1)
    title_cell.font = Font(bold=True, size=16)

    # Fusionner les cellules du titre sur la largeur des en-têtes
    ws.merge_cells(
        start_row=start_row, start_column=1, end_row=start_row, end_column=max(2, len(headers))
    )
    title_cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical='center')

    # Ajouter les en-têtes et les mettre en gras et centrés
    header_row = start_row + 1
    ws.append(headers)

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.font = Font(bold=True)  # Mettre en gras
        cell.alignment = Alignment(horizontal="center", vertical='center')  # Centrer horizontalement

    # Retourner la ligne suivante après le tableau
    return start_row + 2


def merge_groups(ws, start_row, end_row, merge_column, check_column):
    """
    Fusionne les cellules d'une colonne spécifiée si elles contiennent des valeurs identiques
    dans une colonne de vérification spécifiée, sur des lignes consécutives.

    Args:
        ws (openpyxl.worksheet.worksheet.Worksheet): La feuille Excel dans laquelle effectuer la fusion.
        start_row (int): Le numéro de la ligne de début.
        end_row (int): Le numéro de la ligne de fin.
        merge_column (str): La lettre de la colonne dans laquelle fusionner les cellules.
        check_column (str): La lettre de la colonne dans laquelle vérifier les valeurs identiques.
    """
    merge_start_row = start_row
    previous_value = ws[f"{check_column}{start_row}"].value

    for row in range(start_row + 1, end_row + 1):
        current_value = ws[f"{check_column}{row}"].value

        if current_value != previous_value:
            if row - 1 > merge_start_row:
                # Fusionner les cellules de la colonne `merge_column` de la ligne merge_start_row à row-1
                ws.merge_cells(
                    f"{merge_column}{merge_start_row}:{merge_column}{row - 1}"
                )
            # Réinitialiser la valeur de départ pour la prochaine fusion
            merge_start_row = row
            previous_value = current_value

    # Fusionner les dernières cellules si nécessaire
    if end_row > merge_start_row:
        ws.merge_cells(f"{merge_column}{merge_start_row}:{merge_column}{end_row}")

    # Optionnel: Ajuster les alignements des cellules fusionnées
    for row in range(start_row, end_row + 1):
        cell = ws[f"{merge_column}{row}"]
        cell.alignment = cell.alignment.copy(horizontal="center")


def adjust_columns(wb, non_formated_cells):
    """
    Ajuste la largeur de chaque colonne en fonction de la longueur du texte le plus long dans chaque colonne
    et active le retour à la ligne automatique dans toutes les cellules pour toutes les feuilles du classeur.

    :param wb: Le classeur Excel (Workbook)
    """
    for ws in wb.worksheets:
        max_lengths = {}

        # Trouver la longueur maximale du texte pour chaque colonne
        for row in ws.iter_rows():
            for cell in row:
                col_letter = get_column_letter(
                    cell.column
                )  # Utiliser get_column_letter pour obtenir la lettre de la colonne
                if cell.value is not None:
                    cell_length = len(str(cell.value))
                else:
                    cell_length = 0
                if (
                    col_letter not in max_lengths
                    or cell_length > max_lengths[col_letter]
                ):
                    max_lengths[col_letter] = cell_length

        # Ajuster la largeur des colonnes en fonction des longueurs maximales trouvées
        for col_letter, length in max_lengths.items():
            # Utiliser un facteur d'ajustement pour éviter les colonnes trop larges
            if length > 45:
                adjusted_width = 45
            else:
                adjusted_width = length
            ws.column_dimensions[col_letter].width = max(
                10, adjusted_width
            )  # Largeur minimale de 10 pour éviter trop de réduction

        # Activer le retour à la ligne automatique et centrer horizontalement dans toutes les cellules
        for row in ws.iter_rows():
            for cell in row:
                # Vérifier si la cellule n'est pas en gras
                if not (cell.font.bold) and cell.coordinate not in non_formated_cells:
                    cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                


def close_excel_if_open(file_path):
    """
    Vérifie si un fichier Excel est ouvert en recherchant les processus liés à Excel.

    Args:
        file_path (str): Le chemin du fichier à vérifier.

    Returns:
        bool: True si le fichier est ouvert, sinon False.
    """
    file_name = os.path.basename(file_path)

    for proc in psutil.process_iter(["pid", "name"]):
        try:
            if (
                proc.info["name"].lower() in ["excel.exe", "excel"]
                or "EXCEL" in proc.info["name"].upper()
            ):
                for open_file in proc.open_files():
                    if file_name in open_file.path:
                        print(
                            f"Le fichier {file_name} est ouvert. Tentative de fermeture..."
                        )
                        proc.terminate()
                        proc.wait()  # Assure que le processus est bien terminé
                        print(f"Fermeture du fichier {file_name} réussie.")
                        return

        except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
            continue


def apply_borders(wb):
    """
    Applique des bordures intérieures et extérieures aux tableaux présents dans la feuille de calcul,
    en ignorant les lignes vides qui séparent les tableaux.
    
    :param wb: Le classeur (Workbook)
    """
    for ws in wb.worksheets:  # Itérer sur chaque feuille de calcul
        start_row = None  # Pour stocker le début du tableau
        for row in ws.iter_rows():
            # Vérifier si la ligne est vide
            if all(cell.value is None for cell in row):
                # Si on a atteint une ligne vide et qu'un tableau était en cours
                if start_row is not None:
                    # Appliquer les bordures du tableau
                    apply_borders_range(ws, start_row, row[0].row - 1)
                    start_row = None  # Réinitialiser le début du tableau
            else:
                # Si nous sommes dans un tableau, définir le début
                if start_row is None:
                    start_row = row[0].row

        # Gérer le dernier tableau s'il n'est pas suivi d'une ligne vide
        if start_row is not None:
            apply_borders_range(ws, start_row, ws.max_row)  # Corriger ici


def apply_borders_range(ws, start_row, end_row):
    """
    Applique les bordures à la plage de cellules définie par start_row et end_row.
    
    :param ws: La feuille de calcul (Worksheet)
    :param start_row: Ligne de début du tableau
    :param end_row: Ligne de fin du tableau
    """
    # Définir le style des bordures
    thin = Side(border_style="thin", color="000000")  # Bordure fine
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    # Déterminer la première et la dernière colonne contenant des valeurs
    min_col = ws.min_column  # On commence par la première colonne
    max_col = 0  # On va trouver la dernière colonne avec des valeurs

    # Trouver la dernière colonne avec des valeurs dans le tableau
    for row in range(start_row, end_row + 1):
        for col in range(1, ws.max_column + 1):  # Vérifier jusqu'à la dernière colonne de la feuille
            if ws.cell(row=row, column=col).value is not None:
                max_col = max(max_col, col)  # Mettre à jour max_col si on trouve une valeur

    # Appliquer des bordures intérieures
    for row in range(start_row, end_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = Border(left=Side(style='thin'), 
                                 right=Side(style='thin'), 
                                 top=Side(style='thin'), 
                                 bottom=Side(style='thin'))

    # Appliquer bordures extérieures seulement à la première et dernière ligne et colonne
    for col in range(min_col, max_col + 1):
        # Bordure supérieure
        ws.cell(row=start_row, column=col).border = ws.cell(row=start_row, column=col).border + Border(top=Side(style='thin'))
        # Bordure inférieure
        ws.cell(row=end_row, column=col).border = ws.cell(row=end_row, column=col).border + Border(bottom=Side(style='thin'))

    for row in range(start_row, end_row + 1):
        # Bordure gauche
        ws.cell(row=row, column=min_col).border = ws.cell(row=row, column=min_col).border + Border(left=Side(style='thin'))
        # Bordure droite
        ws.cell(row=row, column=max_col).border = ws.cell(row=row, column=max_col).border + Border(right=Side(style='thin'))