import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def extract_general_info(root, tag_paths):
    """Extrait les valeurs des balises spécifiées depuis le XML et retourne une liste des valeurs en string.

    Args:
        root (ET.Element): La racine du document XML.
        tag_paths (list of str): Liste des chemins de balises à extraire.

    Returns:
        list of str: Liste des valeurs extraites sous forme de chaînes de caractères.
    """
    values = []
    for path in tag_paths:
        element = root.find(path)
        # Ajoute la valeur de l'élément ou une chaîne vide si l'élément est None
        values.append(
            element.text if element is not None and element.text is not None else ""
        )
    return values


def create_table(ws, title, headers, start_row):
    """Crée un tableau avec un titre et des en-têtes stylisés dans la feuille de calcul."""
    # Ajouter le titre du tableau et le mettre en gras
    ws.append([title])
    title_cell = ws.cell(row=start_row, column=1)
    title_cell.font = Font(bold=True)

    # Fusionner les cellules du titre sur la largeur des en-têtes
    ws.merge_cells(
        start_row=start_row, start_column=1, end_row=start_row, end_column=len(headers)
    )

    # Ajouter les en-têtes et les mettre en gras et centrés
    header_row = start_row + 1
    ws.append(headers)

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.font = Font(bold=True)  # Mettre en gras
        cell.alignment = Alignment(horizontal="center")  # Centrer horizontalement

    # Retourner la ligne suivante après le tableau
    return start_row + 2


def adjust_columns(wb):
    """
    Ajuste la largeur de chaque colonne en fonction de la longueur du texte le plus long dans chaque colonne
    et active le retour à la ligne automatique dans toutes les cellules pour toutes les feuilles du classeur.

    :param wb: Le classeur Excel (Workbook)
    """
    for ws in wb.worksheets:
        max_lengths = {}

        # Trouver la longueur maximale du texte pour chaque colonne
        for row in ws.iter_rows():
            for cell in row:
                col_letter = get_column_letter(
                    cell.column
                )  # Utiliser get_column_letter pour obtenir la lettre de la colonne
                if cell.value is not None:
                    cell_length = len(str(cell.value))
                else:
                    cell_length = 0
                if (
                    col_letter not in max_lengths
                    or cell_length > max_lengths[col_letter]
                ):
                    max_lengths[col_letter] = cell_length

        # Ajuster la largeur des colonnes en fonction des longueurs maximales trouvées
        for col_letter, length in max_lengths.items():
            # Utiliser un facteur d'ajustement pour éviter les colonnes trop larges
            if length > 45:
                adjusted_width = 45
            else:
                adjusted_width = length
            ws.column_dimensions[col_letter].width = max(
                10, adjusted_width
            )  # Largeur minimale de 10 pour éviter trop de réduction

        # Activer le retour à la ligne automatique dans toutes les cellules
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="center")
