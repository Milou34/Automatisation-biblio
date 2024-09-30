from openpyxl.styles import Alignment, Font

def legende_habitats(ws, current_row, non_formated_cells):
    # Définition des lignes de texte
    lignes_de_texte = [
        "• PF : Forme prioritaire de l'habitat.",
        "• Qualité des données : G = «Bonne» (données reposant sur des enquêtes, par exemple); M = «Moyenne» (données partielles + extrapolations, par exemple); P = «Médiocre» (estimation approximative, par exemple).",
        "• Représentativité : A = «Excellente» ; B = «Bonne» ; C = «Significative» ; D = «Présence non significative».",
        "• Superficie relative : A = 100 ≥ p > 15 % ; B = 15 ≥ p > 2 % ; C = 2 ≥ p > 0 %.",
        "• Conservation : A = «Excellente» ; B = «Bonne» ; C = «Moyenne / réduite».",
        "• Evaluation globale : A = «Excellente» ; B = «Bonne» ; C = «Significative»."
    ]
    
    # Écrire chaque ligne dans une cellule différente
    for ligne in lignes_de_texte:
        cell = ws.cell(row=current_row, column=1)
        cell.value = ligne
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
        cell.font = Font(size=9)
        
        # Ajouter la coordonnée de la cellule à la liste
        non_formated_cells.append(cell.coordinate)
        
        # Incrémenter la ligne actuelle
        current_row += 1
    
    return current_row, non_formated_cells

def legende_especes_inscrites(ws, current_row, non_formated_cells):
    # Définition des lignes de texte
    lignes_de_texte = [
        "• Groupe : A = Amphibiens, B = Oiseaux, F = Poissons, I = Invertébrés, M = Mammifères, P = Plantes, R = Reptiles.",
        "• Type : p = espèce résidente (sédentaire), r = reproduction (migratrice), c = concentration (migratrice), w = hivernage (migratrice).",
        "• Unité : i = individus, p = couples, adults = Adultes matures, area = Superficie en m², bfemales = Femelles reproductrices, cmales = Mâles chanteurs, colonies = Colonies, fstems = Tiges florales, grids1x1 = Grille 1x1 km, grids10x10 = Grille 10x10 km, grids5x5 = Grille 5x5 km, length = Longueur en km, localities = Stations, logs = Nombre de branches, males = Mâles, shoots = Pousses, stones = Cavités rocheuses, subadults = Sub-adultes, trees = Nombre de troncs, tufts = Touffes.",
        "• Catégories du point de vue de l’abondance (Cat.) : C = espèce commune, R = espèce rare, V = espèce très rare, P : espèce présente.",
        "• Qualité des données : G = «Bonne» (données reposant sur des enquêtes, par exemple); M = «Moyenne» (données partielles + extrapolations, par exemple); P = «Médiocre» (estimation approximative, par exemple); DD = Données insuffisantes.",
        "• Population : A = 100 ≥ p > 15 % ; B = 15 ≥ p > 2 % ; C = 2 ≥ p > 0 % ; D = Non significative.",
        "• Conservation : A = «Excellente» ; B = «Bonne» ; C = «Moyenne / réduite».",
        "• Isolement : A = population (presque) isolée ; B = population non isolée, mais en marge de son aire de répartition ; C = population non isolée dans son aire de répartition élargie.",
        "• Evaluation globale : A = «Excellente» ; B = «Bonne» ; C = «Significative»."
    ]
    
    # Écrire chaque ligne dans une cellule différente
    for ligne in lignes_de_texte:
        cell = ws.cell(row=current_row, column=1)
        cell.value = ligne
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
        cell.font = Font(size=9)
        
        # Ajouter la coordonnée de la cellule à la liste
        non_formated_cells.append(cell.coordinate)
        
        # Incrémenter la ligne actuelle
        current_row += 1
    
    return current_row, non_formated_cells


def legende_especes_autres(ws, current_row, non_formated_cells):
    # Définition des lignes de texte
    lignes_de_texte = [
        "• Groupe : A = Amphibiens, B = Oiseaux, F = Poissons, Fu = Champignons, I = Invertébrés, L = Lichens, M = Mammifères, P = Plantes, R = Reptiles.",
        "• Unité : i = individus, p = couples, adults = Adultes matures, area = Superficie en m², bfemales = Femelles reproductrices, cmales = Mâles chanteurs, colonies = Colonies, fstems = Tiges florales, grids1x1 = Grille 1x1 km, grids10x10 = Grille 10x10 km, grids5x5 = Grille 5x5 km, length = Longueur en km, localities = Stations, logs = Nombre de branches, males = Mâles, shoots = Pousses, stones = Cavités rocheuses, subadults = Sub-adultes, trees = Nombre de troncs, tufts = Touffes.",
        "• Catégories du point de vue de l’abondance (Cat.) : C = espèce commune, R = espèce rare, V = espèce très rare, P : espèce présente.",
        "• Motivation : IV, V : annexe où est inscrite l’espèce (directive «Habitats») ; A : liste rouge nationale ; B : espèce endémique ; C : conventions internationales ; D : autres raisons."
    ]
    
    # Écrire chaque ligne dans une cellule différente
    for ligne in lignes_de_texte:
        cell = ws.cell(row=current_row, column=1)
        cell.value = ligne
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
        cell.font = Font(size=9)
        
        # Ajouter la coordonnée de la cellule à la liste
        non_formated_cells.append(cell.coordinate)
        
        # Incrémenter la ligne actuelle
        current_row += 1
    
    return current_row, non_formated_cells