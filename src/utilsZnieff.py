from openpyxl.styles import Font, Alignment

def create_table(ws, title, headers, start_row):
    """Crée un tableau avec un titre et des en-têtes stylisés dans la feuille de calcul."""
    # Ajouter le titre du tableau et le mettre en gras
    ws.append([title])
    title_cell = ws.cell(row=start_row, column=1)
    title_cell.font = Font(bold=True)

    # Fusionner les cellules du titre sur la largeur des en-têtes
    ws.merge_cells(
        start_row=start_row, start_column=1, end_row=start_row, end_column=len(headers)
    )

    # Ajouter les en-têtes et les mettre en gras et centrés
    header_row = start_row + 1
    ws.append(headers)

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.font = Font(bold=True)  # Mettre en gras
        cell.alignment = Alignment(horizontal="center")  # Centrer horizontalement

    # Retourner la ligne suivante après le tableau
    return start_row + 2


