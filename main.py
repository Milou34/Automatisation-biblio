import msvcrt
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font
import fnmatch
import os
from src.n2000.n2000 import process_n2000
from src.utils.utils import adjust_columns, apply_borders, close_excel_if_open, create_table_autres_zones
from src.znieff.znieff import process_znieff
from src.telechargement.telechargementXML import input_telechargement_xml


def main():
        # Validation de l'entrée utilisateur pour le chemin du dossier
    while True:
        folder_source = input_telechargement_xml()
        if folder_source:  # Vérifie que l'entrée n'est pas vide
            break
        print("Erreur : Le dossier source ne peut pas être vide. Veuillez réessayer.")

    non_formated_cells = []
    
        # Validation de l'entrée utilisateur pour le numéro du projet
    while True:
        nom_excel = input("Entrer le numéro du projet : ").strip()  # Assainir l'entrée
        if nom_excel:  # Vérifie que l'entrée n'est pas vide
            break
        print("Erreur : Le numéro du projet ne peut pas être vide. Veuillez réessayer.")

    nom_excel = "Bibliographie-" + str(nom_excel) + ".xlsx"
    output_file = os.path.join(folder_source, nom_excel)

    # Vérifier si le fichier Excel est ouvert et le fermer si c'est le cas
    if close_excel_if_open(output_file):
        print(f"Fermeture du fichier {output_file} réussie.")

    # Créer un nouveau fichier Excel
    wb = Workbook()

    # Créer des feuilles distinctes pour toutes les zones
    ws_donnees_publiques = wb.active
    ws_donnees_publiques.title = "Données publiques"
    ws_znieff1 = wb.create_sheet(title="ZNIEFF 1")
    ws_znieff2 = wb.create_sheet(title="ZNIEFF 2")
    ws_n2000 = wb.create_sheet(title="N2000")
    ws_rnr = wb.create_sheet(title="RNR")
    ws_rnn = wb.create_sheet(title="RNN")
    ws_pnr = wb.create_sheet(title="PNR")
    ws_pnn = wb.create_sheet(title="PNN")
    ws_apb = wb.create_sheet(title="APB")
    ws_rb = wb.create_sheet(title="Réserve biologique")
    ws_mnb = wb.create_sheet(title="Man and Biosphère")
    ws_cen = wb.create_sheet(title="CEN")
    ws_mc = wb.create_sheet(title="Mesures compensatoires")
    
    autres_zones = [ws_rnr, ws_rnn, ws_pnr, ws_pnn, ws_apb, ws_rb, ws_mnb, ws_cen, ws_mc]

    # Initialiser la ligne actuelle pour les ZNIEFF et N2000
    current_row_znieff1 = 1
    current_row_znieff2 = 1
    current_row_n2000 = 1

    # Ajoute les tableaux types pour les autres zones
    ws_donnees_publiques['A1'] = "l’INPN, le MNHN, la DREAL, Tela-botanica, Faune France ainsi que les bases de données naturalistes des associations de la région"
    ws_donnees_publiques['A1'].font = Font(size=16)
    ws_donnees_publiques.merge_cells('A1:N1')
    for ws in autres_zones:
        create_table_autres_zones(ws)

    # Parcours des fichiers dans le dossier
    for chemin, sous_dossiers, fichiers in os.walk(folder_source):
        # Parcourir les fichiers XML trouvés
        for fichier in fichiers:
            if fnmatch.fnmatch(fichier, "*.xml"):
                # Parse le fichier XML
                tree = ET.parse(os.path.join(chemin, fichier))
                root = tree.getroot()
                file_path = os.path.join(chemin, fichier)
                # Sélectionner la feuille et la ligne en fonction du type de ZNIEFF
                if root.tag == "ZNIEFF":
                    # Déterminer le type de ZNIEFF à partir de la balise TY_ZONE
                    type_znieff = int(root.find("TY_ZONE").text)  # On suppose que TY_ZONE existe et est valide
                    if type_znieff == 1:
                        ws = ws_znieff1
                        current_row = process_znieff(ws, root, current_row_znieff1)
                    else:
                        ws = ws_znieff2
                        current_row = process_znieff(ws, root, current_row_znieff2)
                else:
                    ws = ws_n2000
                    current_row, non_formated_cells = process_n2000(ws, root, current_row_n2000, non_formated_cells)

                # Ajoute 2 lignes vides entre chaque fichier XML
                ws.append([])
                ws.append([])
                # Mettre à jour la ligne après avoir ajouté une ligne vide
                current_row += 2

                # Mettre à jour la ligne courante pour le type de ZNIEFF
                if root.tag == "ZNIEFF":
                    type_znieff = int(root.find("TY_ZONE").text)
                    if type_znieff == 1:
                        current_row_znieff1 = current_row
                    else:
                        current_row_znieff2 = current_row
                else:
                    current_row_n2000 = current_row

                # Supprimer le fichier XML après son traitement
                os.remove(file_path)
                print(f"Le fichier {fichier} a été traité puis supprimé.")

    # Sauvegarder le fichier Excel
    adjust_columns(wb, non_formated_cells)
    apply_borders(wb)
    wb.save(os.path.join(folder_source, nom_excel))

    # Ouvrir le fichier Excel généré
    print(f"Ouverture de {output_file} généré avec succès !")
    os.startfile(output_file)

# Lancer le script avec gestion d'erreur
while True:
    try:
        main()
    except Exception as e:
        print(f'Une erreur est arrivée : {e}')
    
    # Attendre une touche de l'utilisateur
    print("\nAppuyez sur 'Entrée' pour relancer le programme, ou une autre touche pour quitter.")
    
    # Lire un caractère sans attendre la touche "Entrée"
    key = msvcrt.getch()
    
    # Si la touche est "Entrée" (caractère 13), relancer la boucle
    if key == b'\r':  # '\r' est le code pour la touche "Entrée" sous Windows
        continue
    else:
        print("Fermeture du programme.")
        break
