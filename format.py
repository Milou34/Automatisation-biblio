import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font
import fnmatch
import os

def xml_to_excel(folder_source):

    # Liste pour stocker les fichiers trouvés
    fichiers_xml = []

    # Parcours des fichiers dans le dossier
    for chemin, sous_dossiers, fichiers in os.walk(folder_source):
        for fichier in fichiers:
            if fnmatch.fnmatch(fichier, '*.xml'):
                fichiers_xml.append(os.path.join(chemin, fichier))

    # Créer un nouveau fichier Excel
    wb = Workbook()
    ws = wb.active
    
    # Affiche les fichiers XML trouvés
    for fichier in fichiers_xml:
        # Parse le fichier XML
        tree = ET.parse(fichier)
        root = tree.getroot()

        # Ajout du titre "Informations générales"
        ws.append(["Informations générales"])

        # Fusionner les cellules pour le titre
        ws.merge_cells('A1:D1')

        # Ajouter les en-têtes du premier tableau
        headers_infos = ['ID national', 'Nom ZNIEFF', 'Type ZNIEFF', 'Surface totale ZNIEFF']
        ws.append(headers_infos)

        # Extraction des données depuis le fichier XML
        id_national = int(root.find('NM_SFFZN').text)
        nom_znieff = root.find('LB_ZN').text
        type_znieff = int(root.find('TY_ZONE').text)
        surface_znieff = float(root.find('SU_ZN').text.replace(',', '.'))
        infos_value = [id_national, nom_znieff, type_znieff, surface_znieff]
        ws.append(infos_value)
        ws.append([])

        # Deuxième tableau avec EUNIS, CORINE, etc.
        ws.append(["Habitats déterminants"])
        ws.merge_cells('A5:F5')

        # Ajouter les en-têtes du second tableau
        habitats_headers = [
            'EUNIS',
            'CORINE biotopes',
            'Habitats d’intérêt communautaire',
            'Source',
            'Observation'
        ]
        ws.append(habitats_headers)
        
        # Parcourir les balises TYPO_INFO_ROW pour récupérer les habitats
        for typo_info_row in root.findall('.//TYPO_INFO_ROW'):
            fg_typo = typo_info_row.find('FG_TYPO').text
            
            # On ne garde que les habitats avec FG_TYPO = "D"
            if fg_typo == "D":
                source = typo_info_row.find('.//AUTEUR')
                source_text = source.text if source.text is not None else ''
                observation_I = typo_info_row.find('.//AN_I_OBS') 
                observation_I_text = observation_I.text if observation_I.text is not None else ''
                observation_S = typo_info_row.find('.//AN_S_OBS') 
                observation_S_text = observation_S.text if observation_S.text is not None else ''
                observation = observation_I_text + " - " + observation_S_text
                
                habitats_values = ["", "", "", source_text, observation]
                
                for typo_row in typo_info_row.findall('.//TYPO_ROW'):
                    lb_typo = typo_row.find('.//LB_TYPO').text
                    lb_hab = typo_row.find('.//LB_HAB').text

                    # Déterminer la colonne selon la valeur de LB_TYPO
                    if lb_typo == "EUNIS 2012":
                        habitats_values[0] = lb_hab
                    elif lb_typo == "CORINE biotopes":
                        habitats_values[1] = lb_hab
                    elif lb_typo == "Habitats d'intérêt communautaire":
                        habitats_values[2] = lb_hab
                        
                print(f"habita value{habitats_values}")
                ws.append(habitats_values)

        # Ajoute une ligne vide entre chaque fichier XML
        ws.append([])

        # Modifier le titre de la feuille en fonction du type ZNIEFF
        if type_znieff == 1:
            ws.title = "ZNIEFF 1"
        elif type_znieff == 2:
            ws.title = "ZNIEFF 2"

    # Sauvegarder le fichier Excel
    wb.save(os.path.join(folder_source, 'Récap.xlsx'))
    print(f"Fichier Excel '{folder_source}/Récap.xlsx' généré avec succès !")

# Demander les chemins des fichiers à l'utilisateur
folder_source = input("Entrez le chemin du dossier des XML sources, le Excel sera créé dans ce dossier : ")

# Lancer la conversion XML vers Excel
xml_to_excel(folder_source)
